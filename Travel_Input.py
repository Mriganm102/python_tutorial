from openpyxl import Workbook

workbook = Workbook()
spreadsheet = workbook.active

spreadsheet["A1"] = "Hello"
spreadsheet["B1"] = "World!"

workbook.save(filename="HelloWorld.xlsx")

from openpyxl import  load_workbook
workbook = load_workbook(filename="HelloWorld.xlsx")
print(workbook.sheetnames)

spreadsheet = workbook.active
print(spreadsheet)
print(spreadsheet["A1"])
print(spreadsheet["A1"].value)

from openpyxl import load_workbook
workbook = load_workbook(filename="HelloWorld.xlsx")
spreadsheet = workbook.active
spreadsheet["C1"]="Hi How are You?"
workbook.save(filename="HelloWorld.xlsx")
print(spreadsheet["C1"])